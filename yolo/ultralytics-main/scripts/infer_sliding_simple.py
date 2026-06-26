"""
原图滑窗推理（简洁版）+ 按缺陷类型分文件夹 + 汇总统计。

流程：滑窗 → 同位置 mask 合并 → NMS → 按 point/line/point_line/no_defect 分目录保存

示例:
  python scripts/infer_sliding_simple.py ^
    --model .../best.pt ^
    --source "H:/卡游/..." ^
    --out "G:/卡游/.../out" ^
    --patch 512 --stride 256 --conf 0.35 ^
    --defect-preprocess none --batch 16
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
import time
from pathlib import Path

import cv2
import numpy as np
import torch

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ultralytics import YOLO
from ultralytics.data.defect_preprocess import apply_defect_preprocess
from ultralytics.utils import ops
from ultralytics.utils.torch_utils import select_device

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
CLASS_COLORS = {0: (0, 255, 0), 1: (0, 0, 255)}
DEFAULT_COLOR = (0, 255, 255)


def imread_unicode(path: Path) -> np.ndarray | None:
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def imwrite_unicode(path: Path, image: np.ndarray) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ext = path.suffix if path.suffix else ".png"
    ok, buf = cv2.imencode(ext, image)
    if not ok:
        raise RuntimeError(f"encode failed: {path}")
    buf.tofile(str(path))


def list_images_here(folder: Path) -> list[Path]:
    return sorted(p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES)


def find_subsets(source: Path) -> list[Path]:
    subsets = []
    for d in sorted(p for p in source.rglob("*") if p.is_dir()):
        if list_images_here(d):
            subsets.append(d)
    if list_images_here(source):
        subsets.insert(0, source)
    return subsets


def unique_target(dst: Path) -> Path:
    if not dst.exists():
        return dst
    stem, suffix = dst.stem, dst.suffix
    idx = 1
    while True:
        candidate = dst.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
        idx += 1


def get_sliding_coords(img_h: int, img_w: int, patch: int, stride: int) -> list[tuple[int, int]]:
    ys = list(range(0, max(img_h - patch, 0) + 1, stride))
    xs = list(range(0, max(img_w - patch, 0) + 1, stride))
    if not ys or ys[-1] + patch < img_h:
        ys.append(max(img_h - patch, 0))
    if not xs or xs[-1] + patch < img_w:
        xs.append(max(img_w - patch, 0))
    return [(y, x) for y in sorted(set(ys)) for x in sorted(set(xs))]


def nms_by_class(dets: list[dict], iou_thr: float) -> list[dict]:
    if not dets:
        return []
    out: list[dict] = []
    by_cls: dict[int, list[dict]] = {}
    for d in dets:
        by_cls.setdefault(d["cls"], []).append(d)

    for items in by_cls.values():
        boxes = np.array([d["xyxy"] for d in items], dtype=np.float32)
        scores = np.array([d["conf"] for d in items], dtype=np.float32)
        idxs = cv2.dnn.NMSBoxes(boxes.tolist(), scores.tolist(), 0.0, iou_thr)
        if len(idxs) == 0:
            continue
        for i in idxs.flatten():
            out.append(items[int(i)])
    return out


def preprocess_patch(patch: np.ndarray, mode: str) -> np.ndarray:
    if mode == "none":
        return patch
    return apply_defect_preprocess(patch, mode)


def extract_patch_mask(result, index: int) -> np.ndarray | None:
    """把模型输出的 mask 映射到 patch 原图坐标（与 boxes.xyxy 同一套 letterbox 逆变换）。

    旧实现用 cv2.resize 强行拉伸，与 scale_boxes 处理的 bbox 坐标系不一致，会导致「有框无 mask」。
    """
    if result.masks is None or len(result.masks) <= index:
        return None
    masks = result.masks.data
    if not isinstance(masks, torch.Tensor):
        masks = torch.as_tensor(masks, dtype=torch.float32)
    else:
        masks = masks.float().cpu()
    if masks.ndim == 2:
        masks = masks.unsqueeze(0)
    # (N, H_lb, W_lb) -> (1, N, H_lb, W_lb) for scale_masks
    scaled = ops.scale_masks(masks[None], result.masks.orig_shape)[0]
    mb = (scaled[index].numpy() > 0.5).astype(np.uint8)
    return mb if mb.any() else None


def xyxy_from_mask(mask: np.ndarray) -> np.ndarray | None:
    ys, xs = np.where(mask > 0)
    if len(xs) == 0:
        return None
    return np.array([xs.min(), ys.min(), xs.max(), ys.max()], dtype=np.float32)


def sync_xyxy_from_mask(det: dict, h: int, w: int) -> None:
    """用 mask 外接矩形覆盖检测头 bbox，保证可视化框与 mask 一致。"""
    mb = det_to_mask(det, h, w)
    xyxy = xyxy_from_mask((mb > 0).astype(np.uint8))
    if xyxy is not None:
        det["xyxy"] = xyxy


def sliding_predict_one(
    model: YOLO,
    img_bgr: np.ndarray,
    patch: int,
    stride: int,
    conf: float,
    imgsz: int,
    defect_preprocess: str,
    batch_size: int,
    names: dict,
    device,
) -> tuple[list[dict], int]:
    h, w = img_bgr.shape[:2]
    coords = get_sliding_coords(h, w, patch, stride)
    all_dets: list[dict] = []

    for start in range(0, len(coords), batch_size):
        chunk = coords[start:start + batch_size]
        patches = [
            preprocess_patch(img_bgr[y0:y0 + patch, x0:x0 + patch].copy(), defect_preprocess)
            for y0, x0 in chunk
        ]
        results = model.predict(patches, imgsz=imgsz, conf=conf, verbose=False, device=device)

        for (y0, x0), r in zip(chunk, results):
            if r.boxes is None or len(r.boxes) == 0:
                continue
            for i in range(len(r.boxes)):
                cls_id = int(r.boxes.cls[i].item())
                score = float(r.boxes.conf[i].item())

                mb = extract_patch_mask(r, i)
                if mb is None:
                    continue

                ph, pw = mb.shape
                mask_full = np.zeros((h, w), np.uint8)
                y2, x2 = min(y0 + ph, h), min(x0 + pw, w)
                mh, mw = y2 - y0, x2 - x0
                mask_full[y0:y2, x0:x2] = mb[:mh, :mw] * 255

                xyxy_local = xyxy_from_mask(mb)
                if xyxy_local is None:
                    continue
                xyxy = xyxy_local.copy()
                xyxy[[0, 2]] += x0
                xyxy[[1, 3]] += y0

                all_dets.append({
                    "xyxy": xyxy,
                    "cls": cls_id,
                    "conf": score,
                    "mask": mask_full,
                    "name": names.get(cls_id, str(cls_id)),
                })

    return all_dets, len(coords)


def box_iou(a: np.ndarray, b: np.ndarray) -> float:
    x1, y1 = max(a[0], b[0]), max(a[1], b[1])
    x2, y2 = min(a[2], b[2]), min(a[3], b[3])
    inter = max(0.0, x2 - x1) * max(0.0, y2 - y1)
    if inter <= 0:
        return 0.0
    aa = (a[2] - a[0]) * (a[3] - a[1])
    ab = (b[2] - b[0]) * (b[3] - b[1])
    return inter / max(aa + ab - inter, 1e-6)


def det_to_mask(det: dict, h: int, w: int) -> np.ndarray:
    mask = det.get("mask")
    if mask is not None and mask.shape[:2] == (h, w):
        return (mask > 0).astype(np.uint8) * 255
    return np.zeros((h, w), np.uint8)


def is_line_det(d: dict) -> bool:
    return "line" in str(d.get("name", "")).lower()


def axis_adjacent(a: dict, b: dict, axis: str, overlap_ratio: float, gap_thr: float) -> bool:
    ax1, ay1, ax2, ay2 = a["xyxy"]
    bx1, by1, bx2, by2 = b["xyxy"]
    if axis == "y":
        cross_ov = max(0.0, min(ax2, bx2) - max(ax1, bx1))
        cross_min = max(1e-6, min(ax2 - ax1, bx2 - bx1))
        along_ov = max(0.0, min(ay2, by2) - max(ay1, by1))
        along_gap = max(0.0, max(ay1, by1) - min(ay2, by2))
        along_span = max(ay2, by2) - min(ay1, by1)
    else:
        cross_ov = max(0.0, min(ay2, by2) - max(ay1, by1))
        cross_min = max(1e-6, min(ay2 - ay1, by2 - by1))
        along_ov = max(0.0, min(ax2, bx2) - max(ax1, bx1))
        along_gap = max(0.0, max(ax1, bx1) - min(ax2, bx2))
        along_span = max(ax2, bx2) - min(ax1, bx1)

    if cross_ov / cross_min < overlap_ratio:
        return False
    return along_ov > 0 or along_gap <= max(gap_thr, 0.15 * along_span)


def dilated_masks_touch(a: dict, b: dict, h: int, w: int, radius: int) -> bool:
    if radius <= 0:
        return False
    k = radius * 2 + 1
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k, k))
    ma = cv2.dilate(det_to_mask(a, h, w), kernel)
    mb = cv2.dilate(det_to_mask(b, h, w), kernel)
    return bool(np.logical_and(ma > 0, mb > 0).any())


def should_merge_dets(
    a: dict,
    b: dict,
    h: int,
    w: int,
    merge_iou: float,
    stride: int = 256,
) -> bool:
    if box_iou(a["xyxy"], b["xyxy"]) >= merge_iou:
        return True

    ma = det_to_mask(a, h, w)
    mb = det_to_mask(b, h, w)
    inter = int(np.logical_and(ma > 0, mb > 0).sum())
    union = int(np.logical_or(ma > 0, mb > 0).sum())
    if union > 0 and inter / union >= merge_iou * 0.5:
        return True

    both_line = is_line_det(a) and is_line_det(b)
    if both_line:
        gap_thr = max(stride * 0.55, 48.0)
        if axis_adjacent(a, b, "y", 0.30, gap_thr):
            return True
        if axis_adjacent(a, b, "x", 0.30, gap_thr):
            return True
    else:
        if axis_adjacent(a, b, "y", 0.50, 20.0):
            return True

    if both_line:
        dilate_r = max(stride // 6, 12)
        if dilated_masks_touch(a, b, h, w, dilate_r):
            return True

    return False


def close_line_mask(mask: np.ndarray, radius: int) -> np.ndarray:
    if radius <= 0 or not mask.any():
        return mask
    k = radius * 2 + 1
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (k, k))
    return cv2.morphologyEx(mask, cv2.MORPH_CLOSE, kernel)


def merge_overlapping_dets(
    dets: list[dict],
    img_shape: tuple[int, ...],
    merge_iou: float,
    stride: int = 256,
) -> list[dict]:
    if not dets:
        return []

    h, w = int(img_shape[0]), int(img_shape[1])
    by_cls: dict[int, list[dict]] = {}
    for d in dets:
        by_cls.setdefault(d["cls"], []).append(d)

    merged: list[dict] = []
    for cls_id, items in by_cls.items():
        used = [False] * len(items)
        for i, seed in enumerate(items):
            if used[i]:
                continue
            group = [seed]
            used[i] = True
            changed = True
            while changed:
                changed = False
                for j, other in enumerate(items):
                    if used[j]:
                        continue
                    for g in group:
                        if should_merge_dets(g, other, h, w, merge_iou, stride=stride):
                            group.append(other)
                            used[j] = True
                            changed = True
                            break

            union_mask = np.zeros((h, w), np.uint8)
            for g in group:
                union_mask = cv2.bitwise_or(union_mask, det_to_mask(g, h, w))

            if is_line_det(group[0]) and len(group) > 1:
                union_mask = close_line_mask(union_mask, max(stride // 8, 8))

            if not union_mask.any():
                continue

            ys, xs = np.where(union_mask > 0)
            xyxy = np.array([xs.min(), ys.min(), xs.max(), ys.max()], dtype=np.float32)

            merged.append({
                "xyxy": xyxy,
                "cls": cls_id,
                "conf": max(g["conf"] for g in group),
                "mask": union_mask,
                "name": group[0]["name"],
            })
    for m in merged:
        sync_xyxy_from_mask(m, h, w)
    return merged


def det_mask_area(det: dict, h: int, w: int) -> int:
    return int(np.count_nonzero(det_to_mask(det, h, w)))


def is_valid_det(det: dict, h: int, w: int, min_mask_area: int) -> bool:
    return det_mask_area(det, h, w) >= min_mask_area


def filter_valid_dets(dets: list[dict], img_shape: tuple[int, ...], min_mask_area: int) -> list[dict]:
    h, w = int(img_shape[0]), int(img_shape[1])
    return [d for d in dets if is_valid_det(d, h, w, min_mask_area)]


def draw_on_full_image(img: np.ndarray, dets: list[dict], mask_alpha: float = 0.35) -> np.ndarray:
    vis = img.copy()
    if vis.ndim == 2:
        vis = cv2.cvtColor(vis, cv2.COLOR_GRAY2BGR)
    overlay = vis.copy()
    h, w = vis.shape[:2]

    for d in dets:
        color = CLASS_COLORS.get(d["cls"], DEFAULT_COLOR)
        label = f"{d['name']} {d['conf']:.2f}"

        mask = det_to_mask(d, h, w)
        if mask.any():
            overlay[mask > 0] = color

        x1, y1, x2, y2 = d["xyxy"].astype(int)
        cv2.rectangle(vis, (x1, y1), (x2, y2), color, 2, cv2.LINE_AA)
        cv2.putText(vis, label, (x1, max(0, y1 - 8)), cv2.FONT_HERSHEY_SIMPLEX, 0.55, color, 2, cv2.LINE_AA)

    if mask_alpha > 0 and dets:
        vis = cv2.addWeighted(overlay, mask_alpha, vis, 1.0 - mask_alpha, 0)
    return vis


def draw_summary_header(
    vis: np.ndarray,
    cls_type: str,
    dets: list[dict],
    raw_n: int,
    after_conf_n: int,
) -> np.ndarray:
    if vis.ndim == 2:
        vis = cv2.cvtColor(vis, cv2.COLOR_GRAY2BGR)
    point_n = sum(1 for d in dets if "point" in str(d["name"]).lower())
    line_n = sum(1 for d in dets if "line" in str(d["name"]).lower())
    text = (
        f"{cls_type}  final={len(dets)} (point={point_n}, line={line_n})  "
        f"raw={raw_n} conf_drop={max(0, raw_n - after_conf_n)} area_drop={max(0, after_conf_n - len(dets))}"
    )
    bar_h = 40
    cv2.rectangle(vis, (0, 0), (vis.shape[1], bar_h), (0, 0, 0), -1)
    cv2.putText(vis, text, (8, 26), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1, cv2.LINE_AA)
    return vis


def finalize_dets(
    dets: list[dict],
    img_shape: tuple[int, ...],
    conf_point: float,
    conf_line: float,
    min_mask_area: int,
) -> tuple[list[dict], int]:
    """统一后处理：置信度过滤 → 有效 mask 过滤。分类与可视化共用此结果。"""
    after_conf = filter_by_class_conf(dets, conf_point, conf_line)
    final = filter_valid_dets(after_conf, img_shape, min_mask_area)
    return final, len(after_conf)


def filter_by_class_conf(dets: list[dict], conf_point: float, conf_line: float) -> list[dict]:
    kept = []
    for d in dets:
        name = str(d["name"]).lower()
        thr = conf_line if "line" in name else conf_point if "point" in name else min(conf_point, conf_line)
        if d["conf"] >= thr:
            kept.append(d)
    return kept


def classify_defect_from_dets(dets: list[dict]) -> str:
    if not dets:
        return "no_defect"
    labels = [str(d["name"]).lower() for d in dets]
    has_point = any("point" in x for x in labels)
    has_line = any("line" in x for x in labels)
    if has_point and has_line:
        return "defect_point_line"
    if has_point:
        return "defect_point"
    if has_line:
        return "defect_line"
    return "defect_point_line"


def save_positive(
    out_dir: Path,
    name: str,
    img_path: Path,
    img: np.ndarray,
    dets: list[dict],
    draw_on_positive: bool,
    mask_alpha: float,
    cls_type: str,
    raw_n: int,
    after_conf_n: int,
) -> None:
    stem = Path(name).stem
    src_suffix = img_path.suffix or ".bmp"

    pred_target = unique_target(out_dir / f"{stem}_pred{src_suffix}")
    if draw_on_positive and dets:
        vis = draw_on_full_image(img, dets, mask_alpha=mask_alpha)
        vis = draw_summary_header(vis, cls_type, dets, raw_n, after_conf_n)
        imwrite_unicode(pred_target, vis)
    else:
        shutil.copy2(img_path, pred_target)

    orig_stem = pred_target.stem.replace("_pred", "_orig", 1)
    orig_target = pred_target.with_name(f"{orig_stem}{src_suffix}")
    shutil.copy2(img_path, orig_target)


def write_summary(rows: list[dict], out_root: Path, summary_name: str) -> None:
    headers = ["大类", "子文件夹", "总数", "point", "line", "point_line", "no_defect"]
    cols = ["group", "subset", "total", "point", "line", "point_line", "no_defect"]

    table: list[list] = []
    grand = {k: 0 for k in ("total", "point", "line", "point_line", "no_defect")}
    last_group = None
    group_acc = {k: 0 for k in grand}

    def flush_group(g):
        if g is None:
            return
        table.append([g, "小计", group_acc["total"], group_acc["point"], group_acc["line"],
                      group_acc["point_line"], group_acc["no_defect"]])

    for row in rows:
        if last_group is not None and row["group"] != last_group:
            flush_group(last_group)
            group_acc = {k: 0 for k in grand}
        table.append([row[c] for c in cols])
        for k in grand:
            grand[k] += row[k]
            group_acc[k] += row[k]
        last_group = row["group"]
    flush_group(last_group)
    table.append(["合计", "ALL", grand["total"], grand["point"], grand["line"],
                  grand["point_line"], grand["no_defect"]])

    csv_path = out_root / (Path(summary_name).stem + ".csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(table)
    print(f"Summary CSV : {csv_path}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "预测汇总"
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="DDDDDD")
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        for r in table:
            ws.append(r)
        sub_fill = PatternFill("solid", fgColor="FFF2CC")
        tot_fill = PatternFill("solid", fgColor="C6EFCE")
        for row_cells in ws.iter_rows(min_row=2):
            label = row_cells[1].value
            if label == "小计":
                for c in row_cells:
                    c.fill = sub_fill
                    c.font = Font(bold=True)
            elif label == "ALL":
                for c in row_cells:
                    c.fill = tot_fill
                    c.font = Font(bold=True)
        widths = [16, 28, 9, 8, 8, 12, 11]
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = wdt
        ws.freeze_panes = "A2"
        xlsx_path = out_root / summary_name
        wb.save(xlsx_path)
        print(f"Summary XLSX: {xlsx_path}")
    except Exception as e:  # noqa: BLE001
        print(f"(xlsx skipped: {e}; CSV is available)")


def process_subset(
    model: YOLO,
    images: list[Path],
    out_base: Path,
    args,
    names: dict,
    device,
    label: str,
) -> dict:
    point_dir = out_base / "defect_point"
    line_dir = out_base / "defect_line"
    point_line_dir = out_base / "defect_point_line"
    no_dir = out_base / "no_defect"
    vis_dir = out_base / "defect_vis"
    for d in (point_dir, line_dir, point_line_dir, no_dir):
        d.mkdir(parents=True, exist_ok=True)
    if args.save_vis:
        vis_dir.mkdir(parents=True, exist_ok=True)
    dropped_dir = out_base / "filtered_out"
    if args.save_dropped:
        dropped_dir.mkdir(parents=True, exist_ok=True)

    conf_point = args.conf if args.conf_point < 0 else args.conf_point
    conf_line = args.conf if args.conf_line < 0 else args.conf_line
    model_conf = min(args.conf, conf_point, conf_line)

    counts = {"point": 0, "line": 0, "point_line": 0, "no_defect": 0}
    detail_rows: list[dict] = []
    total = len(images)

    for done, img_path in enumerate(images, 1):
        img = imread_unicode(img_path)
        if img is None:
            print(f"  Skip unreadable image: {img_path}")
            continue

        t0 = time.time()
        raw_dets, _ = sliding_predict_one(
            model, img, args.patch, args.stride, model_conf, args.imgsz,
            args.defect_preprocess, args.batch, names, device,
        )
        raw_n = len(raw_dets)
        dets = raw_dets
        if not args.no_merge:
            dets = merge_overlapping_dets(dets, img.shape, args.merge_iou, stride=args.stride)
        after_merge_n = len(dets)
        dets = nms_by_class(dets, args.iou)
        after_nms_n = len(dets)
        dets, after_conf_n = finalize_dets(dets, img.shape, conf_point, conf_line, args.min_mask_area)
        elapsed = time.time() - t0

        cls_type = classify_defect_from_dets(dets)
        save_kw = dict(
            cls_type=cls_type,
            raw_n=raw_n,
            after_conf_n=after_conf_n,
        )

        if cls_type == "defect_point":
            save_positive(point_dir, img_path.name, img_path, img, dets, args.draw_on_positive, args.mask_alpha, **save_kw)
            counts["point"] += 1
        elif cls_type == "defect_line":
            save_positive(line_dir, img_path.name, img_path, img, dets, args.draw_on_positive, args.mask_alpha, **save_kw)
            counts["line"] += 1
        elif cls_type == "defect_point_line":
            save_positive(point_line_dir, img_path.name, img_path, img, dets, args.draw_on_positive, args.mask_alpha, **save_kw)
            counts["point_line"] += 1
        else:
            target = unique_target(no_dir / img_path.name)
            shutil.copy2(img_path, target)
            counts["no_defect"] += 1
            if args.save_dropped and raw_n > 0:
                note = f"filtered_out raw={raw_n} merge={after_merge_n} nms={after_nms_n} conf={after_conf_n} final=0"
                vis = draw_summary_header(img.copy(), cls_type, dets, raw_n, after_conf_n)
                imwrite_unicode(unique_target(dropped_dir / f"{img_path.stem}_drop.png"), vis)
                print(f"    dropped: {img_path.name} ({note})")

        if cls_type != "no_defect" and args.save_vis:
            vis = draw_on_full_image(img, dets, mask_alpha=args.mask_alpha)
            vis = draw_summary_header(vis, cls_type, dets, raw_n, after_conf_n)
            imwrite_unicode(unique_target(vis_dir / img_path.name), vis)

        detail_rows.append({
            "filename": img_path.name,
            "cls_type": cls_type,
            "raw": raw_n,
            "after_merge": after_merge_n,
            "after_nms": after_nms_n,
            "after_conf": after_conf_n,
            "final": len(dets),
            "time_sec": round(elapsed, 3),
        })

        print(
            f"  [{label}] [{done}/{total}] {img_path.name} -> {cls_type} "
            f"(raw={raw_n} merge={after_merge_n} nms={after_nms_n} conf={after_conf_n} final={len(dets)} {elapsed:.2f}s) | "
            f"point={counts['point']}, line={counts['line']}, "
            f"point_line={counts['point_line']}, no_defect={counts['no_defect']}"
        )

    detail_path = out_base / "detail.csv"
    if detail_rows:
        with detail_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=list(detail_rows[0].keys()))
            w.writeheader()
            w.writerows(detail_rows)

    counts["total"] = total
    return counts


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Sliding-window inference with defect-type split output")
    p.add_argument("--model", default="H:/Python_cls/YOLO1111111/yolo/runs/segment/runs/seg-weitiao06077/defect_seg/weights/best.pt")
    p.add_argument("--source", default="G:/BaiduNetdiskDownload/新一批图像/好品")
    p.add_argument("--out", default="G:/卡游/haopin")
    p.add_argument("--groups", type=str, nargs="*", default=None, help="Only process subsets whose top-level group is in this list")
    p.add_argument("--summary", type=str, default="预测汇总.xlsx", help="Summary workbook filename written under --out")
    p.add_argument("--patch", type=int, default=512)
    p.add_argument("--stride", type=int, default=256)
    p.add_argument("--imgsz", type=int, default=512)
    p.add_argument("--conf", type=float, default=0.35)
    p.add_argument("--conf-point", type=float, default=-1.0)
    p.add_argument("--conf-line", type=float, default=-1.0)
    p.add_argument("--iou", type=float, default=0.45, help="NMS IoU")
    p.add_argument("--merge-iou", type=float, default=0.30, help="同位置合并 IoU 阈值")
    p.add_argument("--no-merge", action="store_true", help="关闭同位置 mask 合并")
    p.add_argument("--batch", type=int, default=8)
    p.add_argument("--device", default="0")
    p.add_argument("--defect-preprocess", default="none", choices=["none", "point", "line", "mixed"])
    p.add_argument("--mask-alpha", type=float, default=0.35)
    p.add_argument(
        "--min-mask-area",
        type=int,
        default=16,
        help="Minimum mask pixel area; smaller regions are treated as noise and ignored",
    )
    p.add_argument(
        "--save-dropped",
        action="store_true",
        help="Save images that had raw detections but ended up no_defect to filtered_out/",
    )
    p.add_argument(
        "--draw-on-positive",
        action="store_true",
        default=True,
        help="Draw predicted defect regions on images saved to positive folders",
    )
    p.add_argument(
        "--no-draw-on-positive",
        dest="draw_on_positive",
        action="store_false",
        help="Disable drawing on positive images and keep originals",
    )
    p.add_argument("--save-vis", action="store_true", help="Save plotted images to defect_vis")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    out_root = Path(args.out).expanduser().resolve()

    if not source.exists() or not source.is_dir():
        raise FileNotFoundError(f"Source folder not found: {source}")

    subsets = find_subsets(source)
    if args.groups:
        keep = set(args.groups)
        subsets = [s for s in subsets if (s.relative_to(source).parts[:1] or ["."])[0] in keep]
    if not subsets:
        print(f"No image-containing sub-folders found in: {source}")
        return

    print(f"Found {len(subsets)} test subset(s) under {source}:")
    for s in subsets:
        rel = s.relative_to(source)
        print(f"  - {rel if str(rel) != '.' else '<root>'}  ({len(list_images_here(s))} images)")
    print()

    device = select_device(args.device)
    model = YOLO(args.model)
    names = model.names if isinstance(model.names, dict) else dict(enumerate(model.names))

    conf_point = args.conf if args.conf_point < 0 else args.conf_point
    conf_line = args.conf if args.conf_line < 0 else args.conf_line
    model_conf = min(args.conf, conf_point, conf_line)
    print(
        f"[sliding] patch={args.patch} stride={args.stride} conf={model_conf} "
        f"merge_iou={args.merge_iou} merge={not args.no_merge}\n"
    )

    rows: list[dict] = []
    for idx, subset in enumerate(subsets, 1):
        rel = subset.relative_to(source)
        rel_str = "" if str(rel) == "." else str(rel)
        parts = rel.parts
        group = parts[0] if parts and str(rel) != "." else source.name
        sub_name = "\\".join(parts[1:]) if len(parts) > 1 else (parts[0] if parts and str(rel) != "." else "<root>")

        images = list_images_here(subset)
        out_base = out_root / rel if rel_str else out_root
        print(f"==== [{idx}/{len(subsets)}] {group}\\{sub_name}  ({len(images)} images) ====")
        print(f"     out -> {out_base}")

        counts = process_subset(model, images, out_base, args, names, device, label=f"{group}/{sub_name}")
        rows.append({"group": group, "subset": sub_name, **counts})
        print(
            f"     -> point={counts['point']}, line={counts['line']}, "
            f"point_line={counts['point_line']}, no_defect={counts['no_defect']}, total={counts['total']}\n"
        )

    out_root.mkdir(parents=True, exist_ok=True)
    write_summary(rows, out_root, args.summary)

    print("\nAll done.")
    print(f"Source     : {source}")
    print(f"Output root: {out_root}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # pragma: no cover
        print(f"Error: {e}")
        sys.exit(1)
