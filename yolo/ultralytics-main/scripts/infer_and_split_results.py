from __future__ import annotations

import argparse
import itertools
import shutil
import sys
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import cv2
import numpy as np

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))


from ultralytics import YOLO
from ultralytics.data.defect_preprocess import apply_defect_preprocess, mixed_preprocess_batch_torch

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser("Infer images and split by detection result")
    parser.add_argument("--model", type=str, default=r"H:\Python_cls\YOLO1111111\yolo\ultralytics-main\runs\segment\runs\seg-weitiao0622\defect_seg\weights\best.pt", help="Model weights path")
    parser.add_argument("--source", type=str, default="G:\卡游切图\汇总_XY\ceshi", help="Input image folder (may contain multi-level test subsets)")
    parser.add_argument("--out", type=str, default="G:/卡游/xy-0622", help="Output root folder; per-subset outputs mirror the source's relative paths")
    parser.add_argument("--groups", type=str, nargs="*", default=None, help="Only process subsets whose top-level group is in this list (e.g. 骑行 望京)")
    parser.add_argument("--summary", type=str, default="预测汇总.xlsx", help="Summary workbook filename written under --out")
    parser.add_argument("--conf", type=float, default=0.5, help="Confidence threshold")
    parser.add_argument("--iou", type=float, default=0.5, help="IoU threshold")
    parser.add_argument("--imgsz", type=int, default=1024, help="Inference image size")
    parser.add_argument("--device", type=str, default="0", help="Device, e.g. 0 or cpu")
    parser.add_argument("--task", type=str, default="detect", choices=["detect", "segment"], help="Model task type")
    parser.add_argument("--batch", type=int, default=8, help="Batch size for GPU inference")
    parser.add_argument("--workers", type=int, default=4, help="Threads for parallel image read+preprocess")
    parser.add_argument(
        "--defect-preprocess",
        type=str,
        default="none",
        choices=["none", "point", "line", "mixed"],
        help="none=raw image (match ses-dfir training); mixed=legacy OpenCV preprocess",
    )
    parser.add_argument(
        "--preprocess-device",
        type=str,
        default="cuda",
        choices=["cpu", "cuda"],
        help="Where to run mixed preprocess: cpu (threaded) or cuda (GPU batch). Only affects mixed.",
    )
    parser.add_argument(
        "--draw-on-positive",
        action="store_true",
        default=True,
        help="Draw predicted defect regions on images saved to has_result",
    )
    parser.add_argument(
        "--no-draw-on-positive",
        dest="draw_on_positive",
        action="store_false",
        help="Disable drawing on has_result images and keep originals",
    )
    parser.add_argument(
        "--save-vis",
        action="store_true",
        help="Save plotted images for detected results to has_result_vis",
    )
    return parser.parse_args()


def imwrite_unicode(path: Path, image) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ext = path.suffix if path.suffix else ".bmp"
    ok, buf = cv2.imencode(ext, image)
    if not ok:
        raise RuntimeError(f"cv2.imencode failed: {path}")
    buf.tofile(str(path))


def imread_unicode(path: Path):
    data = np.fromfile(str(path), dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def list_images(folder: Path) -> list[Path]:
    return sorted([p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES])


def list_images_here(folder: Path) -> list[Path]:
    """Images located directly inside *folder* (not in sub-folders)."""
    return sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES])


def find_subsets(source: Path) -> list[Path]:
    """Every directory under *source* (incl. source) that directly holds images.

    These are the leaf "test sets". Because parent folders in this dataset hold
    no images directly, this partitions all images with no double counting.
    """
    subsets = []
    for d in sorted(p for p in source.rglob("*") if p.is_dir()):
        if list_images_here(d):
            subsets.append(d)
    if list_images_here(source):
        subsets.insert(0, source)
    return subsets


def has_prediction(result, task: str) -> bool:
    if task == "segment":
        if result.masks is not None:
            try:
                return len(result.masks) > 0
            except TypeError:
                return False
        if result.boxes is not None:
            return len(result.boxes) > 0
        return False
    if result.boxes is None:
        return False
    return len(result.boxes) > 0


def classify_defect_type(result, task: str, names) -> str:
    if not has_prediction(result, task):
        return "no_defect"
    if result.boxes is None or len(result.boxes) == 0:
        return "defect_point_line"

    cls_list = result.boxes.cls.detach().cpu().tolist()
    labels = []
    for cls_id in cls_list:
        name = names.get(int(cls_id), str(int(cls_id)))
        labels.append(str(name).lower())

    has_point = any("point" in x for x in labels)
    has_line = any("line" in x for x in labels)
    if has_point and has_line:
        return "defect_point_line"
    if has_point:
        return "defect_point"
    if has_line:
        return "defect_line"
    return "defect_point_line"


def unique_target(dst: Path) -> Path:
    if not dst.exists():
        return dst
    stem, suffix = dst.stem, dst.suffix
    idx = 1
    while True:
        candidate = dst.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
        idx += 1


def save_positive(out_dir: Path, name: str, img_path: Path, result, draw_on_positive: bool) -> None:
    """Save predicted result (tagged _pred) and the original image (tagged _orig) for comparison."""
    stem = Path(name).stem
    src_suffix = img_path.suffix or ".bmp"

    pred_target = unique_target(out_dir / f"{stem}_pred{src_suffix}")
    if draw_on_positive and result is not None:
        imwrite_unicode(pred_target, result.plot())
    else:
        shutil.copy2(img_path, pred_target)

    orig_stem = pred_target.stem.replace("_pred", "_orig", 1)
    orig_target = pred_target.with_name(f"{orig_stem}{src_suffix}")
    shutil.copy2(img_path, orig_target)


def preprocess_one(img_path: Path, mode: str, defer_gpu: bool):
    """Read and preprocess a single image. Returns (img_path, payload_or_None).

    When defer_gpu is True the raw image is returned as payload so GPU preprocess
    can run later on the main thread (worker threads must not touch the GPU).
    """
    raw = imread_unicode(img_path)
    if raw is None:
        return img_path, None
    if mode == "none" or defer_gpu:
        return img_path, raw
    return img_path, apply_defect_preprocess(raw, mode=mode)


def prefetch_preprocess(images: list[Path], mode: str, pool: ThreadPoolExecutor, max_inflight: int, defer_gpu: bool):
    """Yield (img_path, payload) in order, keeping at most max_inflight preprocess tasks running."""
    futures: deque = deque()
    img_iter = iter(images)
    for _ in range(max_inflight):
        try:
            p = next(img_iter)
        except StopIteration:
            break
        futures.append(pool.submit(preprocess_one, p, mode, defer_gpu))

    while futures:
        fut = futures.popleft()
        try:
            p = next(img_iter)
            futures.append(pool.submit(preprocess_one, p, mode, defer_gpu))
        except StopIteration:
            pass
        yield fut.result()


def chunked(iterable, size: int):
    it = iter(iterable)
    while True:
        chunk = list(itertools.islice(it, size))
        if not chunk:
            return
        yield chunk


def process_subset(model, images: list[Path], out_base: Path, args, label: str) -> dict:
    """Run inference on *images*, split into 4 sub-folders under *out_base*.

    Returns a dict with the per-category counts for this subset.
    """
    point_dir = out_base / "defect_point"
    line_dir = out_base / "defect_line"
    point_line_dir = out_base / "defect_point_line"
    no_dir = out_base / "no_defect"
    vis_dir = out_base / "defect_vis"
    for d in (point_dir, line_dir, point_line_dir, no_dir):
        d.mkdir(parents=True, exist_ok=True)
    if args.save_vis:
        vis_dir.mkdir(parents=True, exist_ok=True)

    counts = {"point": 0, "line": 0, "point_line": 0, "no_defect": 0}
    done = 0
    total = len(images)
    batch_size = max(1, args.batch)
    workers = max(1, args.workers)
    max_inflight = batch_size * 2

    use_gpu_mixed = args.preprocess_device == "cuda" and args.defect_preprocess == "mixed"
    torch_device = "cpu" if args.device == "cpu" else f"cuda:{args.device}"

    with ThreadPoolExecutor(max_workers=workers) as pool:
        prepped = prefetch_preprocess(images, args.defect_preprocess, pool, max_inflight, defer_gpu=use_gpu_mixed)
        for batch in chunked(prepped, batch_size):
            valid = []
            for img_path, payload in batch:
                if payload is None:
                    print(f"  Skip unreadable image: {img_path}")
                    continue
                valid.append((img_path, payload))
            done += len(batch)
            if not valid:
                continue

            if use_gpu_mixed:
                inputs = mixed_preprocess_batch_torch([payload for _, payload in valid], device=torch_device)
            else:
                inputs = [payload for _, payload in valid]
            results = model.predict(
                source=inputs, conf=args.conf, iou=args.iou, imgsz=args.imgsz, device=args.device, verbose=False
            )

            for (img_path, _), r in zip(valid, results):
                cls_type = classify_defect_type(r, args.task, model.names)
                plotted = r.plot() if (cls_type != "no_defect" and args.save_vis) else None

                if cls_type == "defect_point":
                    save_positive(point_dir, img_path.name, img_path, r, args.draw_on_positive)
                    counts["point"] += 1
                elif cls_type == "defect_line":
                    save_positive(line_dir, img_path.name, img_path, r, args.draw_on_positive)
                    counts["line"] += 1
                elif cls_type == "defect_point_line":
                    save_positive(point_line_dir, img_path.name, img_path, r, args.draw_on_positive)
                    counts["point_line"] += 1
                else:
                    target = unique_target(no_dir / img_path.name)
                    shutil.copy2(img_path, target)
                    counts["no_defect"] += 1

                if plotted is not None:
                    imwrite_unicode(unique_target(vis_dir / img_path.name), plotted)

            print(
                f"  [{label}] [{done}/{total}] point={counts['point']}, line={counts['line']}, "
                f"point_line={counts['point_line']}, no_defect={counts['no_defect']}"
            )

    counts["total"] = total
    return counts


def write_summary(rows: list[dict], out_root: Path, summary_name: str) -> None:
    """Write the per-subset summary as an .xlsx (if openpyxl present) and a .csv."""
    headers = ["大类", "子文件夹", "总数", "point", "line", "point_line", "no_defect"]
    cols = ["group", "subset", "total", "point", "line", "point_line", "no_defect"]

    # group subtotals + grand total
    table: list[list] = []
    grand = {k: 0 for k in ("total", "point", "line", "point_line", "no_defect")}
    last_group = None
    group_acc = {k: 0 for k in grand}

    def flush_group(g):
        if g is None:
            return
        table.append([g, "小计", group_acc["total"], group_acc["point"], group_acc["line"],
                      group_acc["point_line"], group_acc["no_defect"]])

    for row in rows:
        if last_group is not None and row["group"] != last_group:
            flush_group(last_group)
            group_acc = {k: 0 for k in grand}
        table.append([row[c] for c in cols])
        for k in grand:
            grand[k] += row[k]
            group_acc[k] += row[k]
        last_group = row["group"]
    flush_group(last_group)
    table.append(["合计", "ALL", grand["total"], grand["point"], grand["line"],
                  grand["point_line"], grand["no_defect"]])

    # CSV (utf-8-sig so Excel shows Chinese correctly)
    import csv
    csv_path = out_root / (Path(summary_name).stem + ".csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(table)
    print(f"Summary CSV : {csv_path}")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "预测汇总"
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="DDDDDD")
        for c in ws[1]:
            c.font = Font(bold=True)
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        for r in table:
            ws.append(r)
        # highlight subtotal / total rows
        sub_fill = PatternFill("solid", fgColor="FFF2CC")
        tot_fill = PatternFill("solid", fgColor="C6EFCE")
        for row_cells in ws.iter_rows(min_row=2):
            label = row_cells[1].value
            if label == "小计":
                for c in row_cells:
                    c.fill = sub_fill
                    c.font = Font(bold=True)
            elif label == "ALL":
                for c in row_cells:
                    c.fill = tot_fill
                    c.font = Font(bold=True)
        widths = [16, 28, 9, 8, 8, 12, 11]
        for i, wdt in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = wdt
        ws.freeze_panes = "A2"
        xlsx_path = out_root / summary_name
        wb.save(xlsx_path)
        print(f"Summary XLSX: {xlsx_path}")
    except Exception as e:  # noqa: BLE001
        print(f"(xlsx skipped: {e}; CSV is available)")


def main() -> None:
    args = parse_args()
    source = Path(args.source).expanduser().resolve()
    out_root = Path(args.out).expanduser().resolve()

    if not source.exists() or not source.is_dir():
        raise FileNotFoundError(f"Source folder not found: {source}")

    subsets = find_subsets(source)
    if args.groups:
        keep = set(args.groups)
        subsets = [s for s in subsets if (s.relative_to(source).parts[:1] or ["."])[0] in keep]
    if not subsets:
        print(f"No image-containing sub-folders found in: {source}")
        return

    print(f"Found {len(subsets)} test subset(s) under {source}:")
    for s in subsets:
        rel = s.relative_to(source)
        print(f"  - {rel if str(rel) != '.' else '<root>'}  ({len(list_images_here(s))} images)")
    print()

    use_gpu_mixed = args.preprocess_device == "cuda" and args.defect_preprocess == "mixed"
    if args.preprocess_device == "cuda" and not use_gpu_mixed:
        print(f"Note: GPU preprocess only supports 'mixed'; falling back to CPU for '{args.defect_preprocess}'.")

    model = YOLO(args.model)

    rows: list[dict] = []
    for idx, subset in enumerate(subsets, 1):
        rel = subset.relative_to(source)
        rel_str = "" if str(rel) == "." else str(rel)
        parts = rel.parts
        group = parts[0] if parts and str(rel) != "." else source.name
        sub_name = "\\".join(parts[1:]) if len(parts) > 1 else (parts[0] if parts and str(rel) != "." else "<root>")

        images = list_images_here(subset)
        out_base = out_root / rel if rel_str else out_root
        print(f"==== [{idx}/{len(subsets)}] {group}\\{sub_name}  ({len(images)} images) ====")
        print(f"     out -> {out_base}")

        counts = process_subset(model, images, out_base, args, label=f"{group}/{sub_name}")
        rows.append({"group": group, "subset": sub_name, **counts})
        print(f"     -> point={counts['point']}, line={counts['line']}, "
              f"point_line={counts['point_line']}, no_defect={counts['no_defect']}, total={counts['total']}\n")

    out_root.mkdir(parents=True, exist_ok=True)
    write_summary(rows, out_root, args.summary)

    print("\nAll done.")
    print(f"Source     : {source}")
    print(f"Output root: {out_root}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:  # pragma: no cover
        print(f"Error: {e}")
        sys.exit(1)
